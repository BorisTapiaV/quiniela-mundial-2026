#!/usr/bin/env python
"""Genera la PLANTILLA EN BLANCO de la quiniela (un solo Excel, bracket-challenge).

Modelo (decisión Boris 2026-06-07): UNA sola planilla, llenada completa antes del
primer partido. El jugador ingresa:
  1. Los 72 marcadores de la fase de grupos (los enfrentamientos son fijos; solo goles).
  2. (auto) El cuadro se arma solo desde SUS grupos: posiciones (con desempates FIFA
     replicados vía CLAVE COMPUESTA, validada == engine.py en 48.000 órdenes) + 8 mejores
     terceros vía la tabla 495 -> su R32.
  3. Cascada de ganadores KO por desplegables DEPENDIENTES hasta el campeón (sin goles KO).
  4. Especiales: goleador / primer eliminado / sorpresa (campeón sale del cuadro).

Salida: quiniela/PLANTILLA_QUINIELA_EN_BLANCO.xlsx
Hojas visibles: Instrucciones · Grupos · Eliminatorias · Especiales.
Hojas ocultas (motor de la planilla): _calc · _t495 · _eq.

NOTA: openpyxl no recalcula fórmulas. El ALGORITMO está probado contra el motor;
la TRANSCRIPCIÓN de fórmulas se verifica estructuralmente. El puntaje real lo da
engine.py al ingerir (autoritativo), así que el puntaje nunca depende de estas fórmulas.
"""
import csv, os, sys
try:
    sys.stdout.reconfigure(encoding='utf-8')
except Exception:
    pass
from openpyxl import Workbook
from openpyxl.utils import get_column_letter as CL
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

HERE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))  # quiniela/
DATA = os.path.join(HERE, 'data')
OUT  = os.path.join(HERE, 'PLANTILLA_QUINIELA_EN_BLANCO.xlsx')

GROUPS = list('ABCDEFGHIJKL')
THIRD_SLOT_MATCH = {'3-ABCDF': 74, '3-CDFGH': 77, '3-CEFHI': 79, '3-EHIJK': 80,
                    '3-BEFIJ': 81, '3-AEHIJ': 82, '3-EFGIJ': 85, '3-DEIJL': 87}
THIRD_SLOTS = [74, 77, 79, 80, 81, 82, 85, 87]  # orden de columnas en _t495

# ---------- carga de datos ----------
def load_equipos():
    eq = {}
    with open(os.path.join(DATA, 'equipos.csv'), encoding='utf-8') as f:
        for r in csv.DictReader(f):
            eq[r['code']] = dict(grupo=r['grupo'], pos=int(r['pos']),
                                 nombre=r['nombre_es'], iso=r['iso_bandera'])
    return eq

def load_fixture():
    rows = []
    with open(os.path.join(DATA, 'fixture.csv'), encoding='utf-8') as f:
        for r in csv.DictReader(f):
            r['match_no'] = int(r['match_no'])
            rows.append(r)
    return rows

def load_t495():
    rows = []
    with open(os.path.join(DATA, 'terceros_495.csv'), encoding='utf-8') as f:
        rd = csv.reader(f)
        header = next(rd)
        for r in rd:
            rows.append(r)
    return header, rows

EQ = load_equipos()
FIX = load_fixture()
T495_HEADER, T495_ROWS = load_t495()

# ---------- estilos ----------
F_TITLE = Font(bold=True, size=16, color='1B4965')
F_H     = Font(bold=True, size=11, color='FFFFFF')
F_GRP   = Font(bold=True, size=12, color='FFFFFF')
F_B     = Font(bold=True)
F_MUT   = Font(size=9, color='888888', italic=True)
FILL_IN   = PatternFill('solid', fgColor='FFF2CC')   # amarillo = input
FILL_H    = PatternFill('solid', fgColor='1B4965')   # azul header
FILL_GRP  = PatternFill('solid', fgColor='2A6F97')
FILL_AUTO = PatternFill('solid', fgColor='EDEDED')   # gris = auto
FILL_WIN  = PatternFill('solid', fgColor='FFE8A3')
CEN = Alignment(horizontal='center', vertical='center')
LEFT = Alignment(horizontal='left', vertical='center')
thin = Side(style='thin', color='CCCCCC')
BORD = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_input(c):
    c.fill = FILL_IN; c.alignment = CEN; c.border = BORD

# ============================================================
wb = Workbook()

# ---------- hoja oculta _eq (code -> nombre, iso) ----------
weq = wb.active; weq.title = '_eq'
weq.append(['code', 'nombre', 'iso'])
for code, v in EQ.items():
    weq.append([code, v['nombre'], v['iso']])
weq.sheet_state = 'hidden'

# ---------- hoja oculta _t495 ----------
wt = wb.create_sheet('_t495')
wt.append(T495_HEADER)            # combo,m74,m77,m79,m80,m81,m82,m85,m87
for r in T495_ROWS:
    wt.append(r)
wt.sheet_state = 'hidden'
# col index (1-based) de cada slot en _t495
T495_COL = {int(h[1:]): i + 1 for i, h in enumerate(T495_HEADER) if h.startswith('m')}

# ============================================================
# ---------- hoja Grupos (inputs + standings) ----------
wg = wb.create_sheet('Grupos')
wg['A1'] = 'QUINIELA MUNDIAL 2026 — Pronóstico'; wg['A1'].font = F_TITLE
wg['A2'] = 'Jugador:'; wg['A2'].font = F_B
wg['B2'].fill = FILL_IN; wg['B2'].border = BORD; wg['B2'].alignment = LEFT
wg['D2'] = 'Fecha límite: 11-jun-2026 (antes del 1er partido)'; wg['D2'].font = F_MUT
wg['A3'] = 'Pon el marcador (goles) de los 72 partidos. Las celdas amarillas son tu input. La tabla de la derecha se ordena sola.'
wg['A3'].font = F_MUT
for w, ww in {'A': 4, 'B': 22, 'C': 5, 'D': 3, 'E': 5, 'F': 22, 'G': 3, 'H': 5, 'I': 20, 'J': 5, 'K': 5, 'L': 5}.items():
    wg.column_dimensions[w].width = ww

# refs de celdas de goles por match_no (en hoja Grupos)
gol_cell = {}   # match_no -> (cellGolLocal, cellGolVisita)
group_matches = {g: [] for g in GROUPS}
for m in FIX:
    if m['fase'] == 'grupos':
        group_matches[m['grupo']].append(m)

row = 5
group_block_row = {}  # g -> primera fila del bloque (fila del header del grupo)
for gi, g in enumerate(GROUPS):
    group_block_row[g] = row
    # header del grupo
    wg.cell(row, 1, f'GRUPO {g}').font = F_GRP
    for c in range(1, 7):
        wg.cell(row, c).fill = FILL_GRP
    # header tabla standings (derecha)
    for j, txt in enumerate(['#', 'Equipo', 'Pts', 'DG', 'GF']):
        cc = wg.cell(row, 8 + j, txt); cc.font = F_H; cc.fill = FILL_H; cc.alignment = CEN
    # sub-header partidos
    hdr = row + 1
    for j, txt in enumerate(['Fecha', 'Local', '', '', 'Visita']):
        cc = wg.cell(hdr, 2 + j, txt); cc.font = F_B; cc.alignment = CEN
    # 6 partidos
    mr = hdr + 1
    for k, m in enumerate(group_matches[g]):
        r = mr + k
        wg.cell(r, 2, m['fecha'][5:])                      # MM-DD
        wg.cell(r, 3, EQ[m['local']]['nombre']).alignment = LEFT
        cl = wg.cell(r, 4); style_input(cl)                # gol local
        cv = wg.cell(r, 5); style_input(cv)                # gol visita
        wg.cell(r, 6, EQ[m['visita']]['nombre']).alignment = LEFT
        gol_cell[m['match_no']] = (f"Grupos!D{r}", f"Grupos!E{r}")
    # tabla standings: 4 filas, posiciones 1..4 referenciando _calc
    base = 2 + gi * 4   # fila base del grupo en _calc
    for p in range(4):
        rr = hdr + 1 + p
        wg.cell(rr, 8, p + 1).alignment = CEN
        # nombre del equipo en posición p+1
        wg.cell(rr, 9, f'=IFERROR(VLOOKUP(INDEX(_calc!$B${base}:$B${base+3},'
                       f'MATCH({p+1},_calc!$L${base}:$L${base+3},0)),_eq!$A:$B,2,FALSE),"")')
        wg.cell(rr, 10, f'=IFERROR(INDEX(_calc!$D${base}:$D${base+3},MATCH({p+1},_calc!$L${base}:$L${base+3},0)),"")').alignment = CEN
        wg.cell(rr, 11, f'=IFERROR(INDEX(_calc!$G${base}:$G${base+3},MATCH({p+1},_calc!$L${base}:$L${base+3},0)),"")').alignment = CEN
        wg.cell(rr, 12, f'=IFERROR(INDEX(_calc!$E${base}:$E${base+3},MATCH({p+1},_calc!$L${base}:$L${base+3},0)),"")').alignment = CEN
        for cc in range(8, 13):
            wg.cell(rr, cc).fill = FILL_AUTO; wg.cell(rr, cc).border = BORD
    row = hdr + 1 + 6 + 1   # siguiente bloque (gap de 1)

# DV goles 0..20
dv_goal = DataValidation(type='whole', operator='between', formula1='0', formula2='20', allow_blank=True)
dv_goal.error = 'Pon un número de goles entre 0 y 20'
wg.add_data_validation(dv_goal)
for mn, (cl, cv) in gol_cell.items():
    dv_goal.add(cl.split('!')[1]); dv_goal.add(cv.split('!')[1])
wg.sheet_view.showGridLines = False

# ============================================================
# ---------- hoja oculta _calc (motor de la planilla) ----------
wc = wb.create_sheet('_calc')
wc.sheet_state = 'hidden'
# team -> fila en _calc ; orden por pos dentro del grupo
team_row = {}
group_base = {}
for gi, g in enumerate(GROUPS):
    base = 2 + gi * 4
    group_base[g] = base
    teams = sorted([c for c, v in EQ.items() if v['grupo'] == g], key=lambda c: EQ[c]['pos'])
    for i, t in enumerate(teams):
        team_row[t] = base + i

# helper: matches por equipo (golFor, golAg cells, opp)
def team_matches(t):
    out = []
    g = EQ[t]['grupo']
    for m in group_matches[g]:
        cl, cv = gol_cell[m['match_no']]
        if m['local'] == t:
            out.append((cl, cv, m['visita']))   # golFor, golAg, opp
        elif m['visita'] == t:
            out.append((cv, cl, m['local']))
    return out

for gi, g in enumerate(GROUPS):
    base = group_base[g]
    teams = sorted([c for c, v in EQ.items() if v['grupo'] == g], key=lambda c: EQ[c]['pos'])
    for i, t in enumerate(teams):
        r = base + i
        tm = team_matches(t)
        wc.cell(r, 1, g)
        wc.cell(r, 2, t)
        wc.cell(r, 3, EQ[t]['pos'])                                  # C seed
        # solo cuenta partidos CON marcador ingresado (ambas celdas numéricas);
        # planilla vacía -> tabla en 0 (evita el "3 puntos base" de leer celdas vacías como 0-0)
        def pl(a, b):
            return f'AND(ISNUMBER({a}),ISNUMBER({b}))'
        pts = '+'.join(f'IF({pl(a,b)},IF({a}>{b},3,IF({a}={b},1,0)),0)' for a, b, o in tm)
        gf  = [f'IF({pl(a,b)},{a},0)' for a, b, o in tm]
        ga  = [f'IF({pl(a,b)},{b},0)' for a, b, o in tm]
        wc.cell(r, 4, f'={pts}')                                     # D pts
        wc.cell(r, 5, '=' + '+'.join(gf))                            # E gf
        wc.cell(r, 6, '=' + '+'.join(ga))                            # F ga
        wc.cell(r, 7, f'=E{r}-F{r}')                                 # G gd
        # H hp / I hgd / J hgf  (restringido a oponentes con MISMOS puntos globales, y partido jugado)
        hp, hgd, hgf = [], [], []
        for a, b, o in tm:
            orow = team_row[o]
            cond = f'AND({pl(a,b)},D{orow}=D{r})'
            hp.append(f'IF({cond},IF({a}>{b},3,IF({a}={b},1,0)),0)')
            hgd.append(f'IF({cond},{a}-{b},0)')
            hgf.append(f'IF({cond},{a},0)')
        wc.cell(r, 8, '=' + '+'.join(hp))
        wc.cell(r, 9, '=' + '+'.join(hgd))
        wc.cell(r, 10, '=' + '+'.join(hgf))
        # K key compuesta (lexicográfica, validada == engine)
        wc.cell(r, 11, f'=D{r}*100000000000+H{r}*1000000000+(I{r}+50)*10000000'
                       f'+J{r}*100000+(G{r}+50)*1000+E{r}*10+(4-C{r})')
        # L rank dentro del grupo
        wc.cell(r, 12, f'=RANK(K{r},K{base}:K{base+3},0)')
    # N/O/P = 1º/2º/3º code (en fila base)
    wc.cell(base, 14, f'=INDEX(B{base}:B{base+3},MATCH(1,L{base}:L{base+3},0))')  # N 1º
    wc.cell(base, 15, f'=INDEX(B{base}:B{base+3},MATCH(2,L{base}:L{base+3},0))')  # O 2º
    wc.cell(base, 16, f'=INDEX(B{base}:B{base+3},MATCH(3,L{base}:L{base+3},0))')  # P 3º

# ---- tabla de terceros (filas 52..63, una por grupo en orden A..L) ----
TBASE = 52
for gi, g in enumerate(GROUPS):
    r = TBASE + gi
    base = group_base[g]
    wc.cell(r, 1, g)                                                            # A grupo
    wc.cell(r, 2, f'=P{base}')                                                  # B 3º code
    wc.cell(r, 3, f'=INDEX(D{base}:D{base+3},MATCH(3,L{base}:L{base+3},0))')    # C pts3
    wc.cell(r, 4, f'=INDEX(G{base}:G{base+3},MATCH(3,L{base}:L{base+3},0))')    # D gd3
    wc.cell(r, 5, f'=INDEX(E{base}:E{base+3},MATCH(3,L{base}:L{base+3},0))')    # E gf3
    # F tkey: pts*1e7 + (gd+50)*1e4 + gf*100 + letterord (A=1..L=12, mayor letra gana empate)
    # magnitudes separadas: gf*100 > letterord(<=12); evita solapamiento del desempate
    wc.cell(r, 6, f'=C{r}*10000000+(D{r}+50)*10000+E{r}*100+{gi+1}')
    # G trank (1..12)
    wc.cell(r, 7, f'=RANK(F{r},F{TBASE}:F{TBASE+11},0)')

# ---- combo string (8 mejores terceros, en orden A..L => ya ordenado) ----
combo_parts = []
for gi, g in enumerate(GROUPS):
    r = TBASE + gi
    combo_parts.append(f'IF(G{r}<=8,"{g}","")')
wc.cell(66, 1, '=' + '&'.join(combo_parts))    # A66 = combo
COMBO = '_calc!$A$66'

# ---- código del 3º que juega cada third-slot (filas 70..77) ----
slot_code_row = {}
for i, mn in enumerate(THIRD_SLOTS):
    r = 70 + i
    slot_code_row[mn] = r
    col_idx = T495_COL[mn]
    # letra de grupo cuyo 3º juega ese slot
    wc.cell(r, 1, mn)
    wc.cell(r, 2, f'=VLOOKUP({COMBO},_t495!$A:$I,{col_idx},FALSE)')             # B group letter
    # code del 3º de esa letra
    wc.cell(r, 3, f'=INDEX($B${TBASE}:$B${TBASE+11},MATCH(B{r},$A${TBASE}:$A${TBASE+11},0))')  # C code

def slot_third_code_ref(mn):
    return f'_calc!$C${slot_code_row[mn]}'

# ---- R32 resuelto (filas 80..95, partidos 73..88) ----
def resolve_slot_ref(slot):
    """Devuelve una ref de celda (_calc) con el código del equipo del slot."""
    if slot.startswith('3-'):
        return slot_third_code_ref(THIRD_SLOT_MATCH[slot])
    pos = int(slot[0]); g = slot[1]
    base = group_base[g]
    col = {1: 14, 2: 15, 3: 16}[pos]   # N/O/P
    return f'_calc!${CL(col)}${base}'

r32_row = {}
rr = 80
for m in FIX:
    if m['fase'] == 'R32':
        mn = m['match_no']
        r32_row[mn] = rr
        wc.cell(rr, 1, mn)
        wc.cell(rr, 2, f'={resolve_slot_ref(m["local"])}')    # B local code
        wc.cell(rr, 3, f'={resolve_slot_ref(m["visita"])}')   # C visita code
        rr += 1

# ============================================================
# ---------- hoja Eliminatorias (cuadro + dropdowns dependientes) ----------
we = wb.create_sheet('Eliminatorias')
we['A1'] = 'ELIMINATORIAS — pica quién avanza en cada llave'; we['A1'].font = F_TITLE
we['A2'] = ('Tu cuadro se arma solo desde tus grupos (1º/2º + 8 mejores terceros). '
            'Elige el GANADOR de cada llave; la siguiente ronda se llena con tus ganadores. Sin goles aquí.')
we['A2'].font = F_MUT
for w, ww in {'A': 8, 'B': 6, 'C': 22, 'D': 3, 'E': 22, 'F': 22, 'G': 3, 'H': 8, 'I': 8}.items():
    we.column_dimensions[w].width = ww
we.sheet_view.showGridLines = False

# nombre desde code
def name_formula(code_ref):
    return f'=IFERROR(VLOOKUP({code_ref},_eq!$A:$B,2,FALSE),"")'

ROUNDS = [('R32', 'OCTAVOS DE FINAL (R32)', 73, 88),
          ('R16', 'DIECISÉIS (R16)', 89, 96),
          ('QF',  'CUARTOS', 97, 100),
          ('SF',  'SEMIFINALES', 101, 102),
          ('3P',  'TERCER LUGAR', 103, 103),
          ('Final', 'FINAL', 104, 104)]

fixture_by_no = {m['match_no']: m for m in FIX}
win_cell = {}    # match_no -> celda E del ganador (en Eliminatorias)
loc_cell = {}    # match_no -> celda H (code local)
vis_cell = {}    # match_no -> celda I (code visita)
dvs = []
r = 4
for key, label, lo, hi in ROUNDS:
    cc = we.cell(r, 1, label); cc.font = F_GRP; cc.fill = FILL_GRP
    for k in range(1, 7):
        we.cell(r, k).fill = FILL_GRP
    r += 1
    we.cell(r, 2, '#').font = F_B
    we.cell(r, 3, 'Local').font = F_B
    we.cell(r, 5, 'Visita').font = F_B
    we.cell(r, 6, 'GANADOR').font = F_B
    r += 1
    for mn in range(lo, hi + 1):
        m = fixture_by_no[mn]
        hcell = f'H{r}'; icell = f'I{r}'
        loc_cell[mn] = hcell; vis_cell[mn] = icell
        # NOMBRES completos local/visita (ocultos H/I) -> el dropdown muestra nombres, no códigos
        if key == 'R32':
            we[hcell] = name_formula(f'_calc!$B${r32_row[mn]}')
            we[icell] = name_formula(f'_calc!$C${r32_row[mn]}')
        elif key == '3P':
            # perdedores de las semis (101,102): el que no ganó (comparación por nombre)
            f1 = int(m['local'][1:]); f2 = int(m['visita'][1:])   # L101, L102
            we[hcell] = f'=IF({win_cell[f1]}={loc_cell[f1]},{vis_cell[f1]},{loc_cell[f1]})'
            we[icell] = f'=IF({win_cell[f2]}={loc_cell[f2]},{vis_cell[f2]},{loc_cell[f2]})'
        else:
            f1 = int(m['local'][1:]); f2 = int(m['visita'][1:])   # W74, W77...
            we[hcell] = f'={win_cell[f1]}'
            we[icell] = f'={win_cell[f2]}'
        we[hcell].font = F_MUT; we[icell].font = F_MUT
        # display (los nombres ya están en H/I)
        we.cell(r, 2, mn).alignment = CEN
        we.cell(r, 3, f'={hcell}').alignment = LEFT
        we.cell(r, 4, 'vs').alignment = CEN
        we.cell(r, 5, f'={icell}').alignment = LEFT
        # ganador input (dropdown = los 2 NOMBRES en H:I)
        gc = we.cell(r, 6); style_input(gc); gc.fill = FILL_WIN
        win_cell[mn] = f'F{r}'
        dv = DataValidation(type='list', formula1=f'=$H${r}:$I${r}', allow_blank=True)
        dv.error = 'Elige uno de los dos equipos de la llave'
        we.add_data_validation(dv); dv.add(f'F{r}')
        r += 1
    r += 1
# ocultar columnas helper H, I
we.column_dimensions['H'].hidden = True
we.column_dimensions['I'].hidden = True

# ============================================================
# ---------- hoja Especiales ----------
ws = wb.create_sheet('Especiales')
ws['A1'] = 'ESPECIALES DE TORNEO'; ws['A1'].font = F_TITLE
ws['A2'] = 'El campeón sale de tu cuadro (automático). Completa los demás.'; ws['A2'].font = F_MUT
ws.column_dimensions['A'].width = 26
ws.column_dimensions['B'].width = 26
ws.sheet_view.showGridLines = False

champ_ref = win_cell[104]
fin1 = loc_cell[104]; fin2 = vis_cell[104]
# campeón/subcampeón salen del cuadro y YA son nombres (H/I y ganadores guardan nombres)
rows_sp = [
    ('Campeón (auto)', f'=IF(Eliminatorias!{champ_ref}="","(elige la final)",Eliminatorias!{champ_ref})', 'auto'),
    ('Subcampeón (auto)', f'=IF(Eliminatorias!{champ_ref}="","",'
                          f'IF(Eliminatorias!{champ_ref}=Eliminatorias!{fin1},'
                          f'Eliminatorias!{fin2},Eliminatorias!{fin1}))', 'auto'),
    ('Goleador (Bota de Oro)', '', 'text'),
    ('Primer eliminado', '', 'team'),
    ('Sorpresa del torneo', '', 'text'),
]
r = 4
dv_team = DataValidation(type='list', formula1='=_eq!$B$2:$B$49', allow_blank=True)
ws.add_data_validation(dv_team)
sp_cell = {}
for label, formula, kind in rows_sp:
    ws.cell(r, 1, label).font = F_B
    c = ws.cell(r, 2)
    if kind == 'auto':
        c.value = formula; c.fill = FILL_AUTO; c.border = BORD
    else:
        style_input(c); c.alignment = LEFT
        if kind == 'team':
            dv_team.add(f'B{r}')
            ws.cell(r, 3, '(elige de la lista)').font = F_MUT
        else:
            ws.cell(r, 3, '(texto libre)').font = F_MUT
    sp_cell[label] = f'B{r}'
    r += 1

# ============================================================
# ---------- hoja Instrucciones (primera) ----------
wi = wb.create_sheet('Instrucciones', 0)
wi.sheet_view.showGridLines = False
wi.column_dimensions['A'].width = 100
lines = [
    ('QUINIELA MUNDIAL 2026 — Cómo llenar tu pronóstico', F_TITLE),
    ('', None),
    ('Fecha límite: 11 de junio de 2026, antes del primer partido.', F_B),
    ('', None),
    ('1) Hoja GRUPOS — pon el marcador (goles) de los 72 partidos de la fase de grupos.', F_B),
    ('   Solo escribes goles en las celdas amarillas; los equipos ya están fijos.', None),
    ('   La tabla de la derecha de cada grupo se ordena sola con los desempates oficiales FIFA.', None),
    ('', None),
    ('2) Hoja ELIMINATORIAS — tu cuadro se arma solo desde tus grupos (1º, 2º y los 8 mejores terceros).', F_B),
    ('   Elige el GANADOR de cada llave en la lista desplegable. La ronda siguiente se llena con tus ganadores.', None),
    ('   Aquí no se ponen goles, solo quién avanza, hasta el campeón.', None),
    ('', None),
    ('3) Hoja ESPECIALES — el campeón sale solo de tu cuadro. Completa goleador, primer eliminado y sorpresa.', F_B),
    ('', None),
    ('CÓMO SE PUNTÚA (máximo 718 puntos):', F_B),
    ('   • Grupos (50%): marcador exacto 5 · solo diferencia 3 · solo resultado (1X2) 2. Te quedas con el más alto.', None),
    ('   • Avance KO (35%): por cada equipo tuyo que llega a la ronda — R32 2 · R16 4 · Cuartos 6 · Semis 10 · Final 16.', None),
    ('   • Especiales (15%): campeón 50 · goleador 25 · primer eliminado 20 · sorpresa 15.', None),
    ('   Desempate del ranking: más puntos → más marcadores exactos → acertó campeón.', None),
    ('', None),
    ('Devuelve este mismo archivo lleno. ¡Suerte!', F_MUT),
]
for i, (txt, font) in enumerate(lines, 1):
    c = wi.cell(i, 1, txt)
    if font:
        c.font = font

# orden de hojas visibles
wb.move_sheet('Grupos', -(wb.sheetnames.index('Grupos') - 1))
order = ['Instrucciones', 'Grupos', 'Eliminatorias', 'Especiales', '_calc', '_t495', '_eq']
wb._sheets.sort(key=lambda s: order.index(s.title))

wb.active = wb.sheetnames.index('Instrucciones')
wb.save(OUT)
print('OK ->', OUT)
print('Hojas:', wb.sheetnames)
print(f'Inputs goles: {len(gol_cell)} partidos · llaves KO: {len(win_cell)} · especiales input: 3')
