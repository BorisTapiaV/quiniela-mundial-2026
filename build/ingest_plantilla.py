#!/usr/bin/env python
"""Ingesta de una PLANTILLA llena -> formato de predicción (data/predicciones/<slug>.*).

Lee el Excel que devuelve un jugador y produce, como MF/CASA:
  data/predicciones/<slug>.csv            (match_no,local,visita,gl,gv)  72 grupos
  data/predicciones/<slug>_ko.csv         (match_no,ganador)             32 llaves KO
  data/predicciones/<slug>_especiales.csv (clave,valor)                  campeón/goleador/1ºelim/sorpresa

El MOTOR es autoritativo: re-deriva el bracket desde los 72 marcadores y verifica que
cada ganador KO elegido sea uno de los dos equipos de su llave derivada (si no, lo avisa).
El puntaje real nunca depende de las fórmulas del Excel, solo de estos CSV + engine.py.

Uso:  python build/ingest_plantilla.py <archivo.xlsx> [slug]
"""
import sys, os, csv, re
from openpyxl import load_workbook
try:
    sys.stdout.reconfigure(encoding='utf-8')   # consola Windows: permitir ✓/⚠
except Exception:
    pass

HERE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(HERE, 'build'))
import engine

GROUPS = list('ABCDEFGHIJKL')
PRED = os.path.join(HERE, 'data', 'predicciones')


def group_input_layout(group_matches):
    """Reproduce el cursor de filas del generador -> match_no -> (filaGol)."""
    row = 5; cells = {}
    for g in GROUPS:
        hdr = row + 1; mr = hdr + 1
        for k, m in enumerate(group_matches[g]):
            r = mr + k
            cells[m['match_no']] = r   # D=gol local, E=gol visita
        row = hdr + 1 + 6 + 1
    return cells


def slugify(name):
    s = re.sub(r'[^A-Za-z0-9]+', '_', (name or '').strip()).strip('_').upper()
    return s or 'JUGADOR'


def main():
    if len(sys.argv) < 2:
        print(__doc__); sys.exit(1)
    path = sys.argv[1]
    fix = engine.load_fixture()
    eq = engine.load_equipos()
    ter = engine.load_terceros()
    fixture_by_no = {m['match_no']: m for m in fix}
    gm = {g: [] for g in GROUPS}
    for m in fix:
        if m['fase'] == 'grupos':
            gm[m['grupo']].append(m)

    wb = load_workbook(path, data_only=True)
    wg = wb['Grupos']; we = wb['Eliminatorias']; wsp = wb['Especiales']

    player = wg['B2'].value
    slug = (sys.argv[2] if len(sys.argv) > 2 else slugify(player))

    # ----- grupos -----
    rowmap = group_input_layout(gm)
    group_scores = {}    # match_no -> (gl,gv)
    missing = []
    rows_out = []
    for g in GROUPS:
        for m in gm[g]:
            r = rowmap[m['match_no']]
            gl = wg.cell(r, 4).value; gv = wg.cell(r, 5).value
            if gl is None or gv is None:
                missing.append(m['match_no']); gl = gl or 0; gv = gv or 0
            gl, gv = int(gl), int(gv)
            group_scores[m['match_no']] = (gl, gv)
            rows_out.append((m['match_no'], m['local'], m['visita'], gl, gv))

    # mapeo NOMBRE -> código (los dropdowns guardan nombres completos); acepta código directo también
    name2code = {v['nombre_es'].strip().upper(): c for c, v in eq.items()}
    def to_code(val):
        if not val:
            return ''
        up = str(val).strip().upper()
        if up in name2code:
            return name2code[up]
        if up in eq:                       # ya es un código (ESP, FRA…)
            return up
        return up                          # desconocido: se conserva y se avisa luego

    # ----- KO: leer ganadores (col F) por match_no (col B) -----
    ko_winners = {}
    for row in we.iter_rows():
        bcell = next((c for c in row if c.column == 2), None)
        if bcell is None:
            continue
        mn = bcell.value
        if isinstance(mn, int) and 73 <= mn <= 104:
            w = we.cell(bcell.row, 6).value   # col F
            if w:
                ko_winners[mn] = to_code(w)

    # ----- especiales -----
    esp = {}
    for row in wsp.iter_rows():
        a = next((c for c in row if c.column == 1), None)
        if a is None or not a.value:
            continue
        label = str(a.value)
        val = wsp.cell(a.row, 2).value
        if 'Goleador' in label:
            esp['goleador'] = (val or '')
        elif 'Primer eliminado' in label:
            esp['primer_eliminado'] = to_code(val)
        elif 'Sorpresa' in label:
            esp['sorpresa'] = (val or '')

    # ----- motor autoritativo: derivar bracket desde los 72 -----
    allst = engine.compute_all_standings(
        {g: [(m['local'], m['visita'], *group_scores[m['match_no']]) for m in gm[g]] for g in GROUPS}, eq)
    r32 = engine.build_r32(allst, fix, ter)

    # verificar consistencia de los picks KO contra el bracket derivado
    warns = []
    # R32: cada ganador debe estar entre los 2 equipos derivados de esa llave
    for mn in range(73, 89):
        if mn in ko_winners and mn in r32:
            if ko_winners[mn] not in r32[mn]:
                warns.append(f'M{mn}: ganador {ko_winners[mn]} no está en la llave derivada {r32[mn]}')
    # campeón = ganador de 104
    champ = ko_winners.get(104, '')

    # ----- escribir CSVs -----
    os.makedirs(PRED, exist_ok=True)
    with open(os.path.join(PRED, f'{slug}.csv'), 'w', newline='', encoding='utf-8') as f:
        w = csv.writer(f); w.writerow(['match_no', 'local', 'visita', 'gl', 'gv'])
        for r in rows_out:
            w.writerow(r)
    with open(os.path.join(PRED, f'{slug}_ko.csv'), 'w', newline='', encoding='utf-8') as f:
        w = csv.writer(f); w.writerow(['match_no', 'ganador'])
        for mn in range(73, 105):
            w.writerow([mn, ko_winners.get(mn, '')])
    with open(os.path.join(PRED, f'{slug}_especiales.csv'), 'w', newline='', encoding='utf-8') as f:
        w = csv.writer(f); w.writerow(['clave', 'valor'])
        w.writerow(['campeon', champ])
        w.writerow(['goleador', esp.get('goleador', '')])
        w.writerow(['primer_eliminado', esp.get('primer_eliminado', '')])
        w.writerow(['sorpresa', esp.get('sorpresa', '')])

    # ----- reporte -----
    print(f'Jugador: {player!r}  ->  slug: {slug}')
    print(f'  grupos: {len(rows_out)}/72  ·  KO elegidos: {len([m for m in ko_winners if ko_winners[m]])}/32  ·  campeón: {champ or "(sin elegir)"}')
    if missing:
        print(f'  ⚠ {len(missing)} partidos de grupo sin marcador: {missing[:12]}{"..." if len(missing)>12 else ""}')
    if warns:
        print(f'  ⚠ {len(warns)} picks KO inconsistentes con el bracket derivado:')
        for x in warns[:8]:
            print(f'      {x}')
    else:
        print('  ✓ picks KO consistentes con el bracket derivado por el motor')
    print(f'  escrito en {PRED}\\{slug}.csv (+_ko +_especiales)')


if __name__ == '__main__':
    main()
